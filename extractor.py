import pandas as pd
import openpyxl
import io
import os
import requests
import json
import re
from urllib.parse import quote
from msal import ConfidentialClientApplication
import google.generativeai as genai
from typing import List, Dict, Optional, Set

class GeminiCompanyExtractor:
    def __init__(self, source, api_key: str, target_company: str = None, max_sheets: int = 10):
        """
        Initialize the extractor with Gemini integration.
        :param source: file path (str) or BytesIO stream (from SharePoint)
        :param api_key: Google Gemini API Key
        :param target_company: Name of the target company for context-aware extraction
        :param max_sheets: number of sheets to process
        """
        self.source = source
        self.target_company = target_company
        self.max_sheets = max_sheets
        self.results = {}
        
        # Configure Gemini
        genai.configure(api_key=api_key)

    def _convert_to_dataframe(self, workbook):
        """
        Convert Excel workbook to DataFrame, prioritizing sheets with 'comps' variations.
        """
        all_data = []
        sheet_data = {}

        # 1. Identify the best sheet(s) to process
        target_sheet_names = []
        sheet_names = workbook.sheetnames
        
        # Regex pattern for flexible matching:
        # Matches "equity", "trading", "public" followed by anything (or nothing) then "comps"
        # OR just "comps"
        # Handles separators like space, underscore, dash, etc.
        pattern = re.compile(r'(equity|trading|public).*comps|comps', re.IGNORECASE)
        
        for name in sheet_names:
            if pattern.search(name):
                target_sheet_names.append(name)
                # If we found a very specific match, we might stop, but let's collect all candidates
                # and maybe just pick the first one or all of them? 
                # User said "one sheet found... matching the name". 
                # Let's pick the first match as the primary target.
                break
        
        if target_sheet_names:
            print(f"✅ Found target sheet: {target_sheet_names[0]}")
        else:
            # Fallback: Use first 3 sheets
            target_sheet_names = sheet_names[:3]
            print(f"⚠️ No 'comps' sheet found. Defaulting to first {len(target_sheet_names)} sheets: {target_sheet_names}")

        # 2. Process the identified sheet(s)
        for sheet_name in target_sheet_names:
            sheet = workbook[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
                # Convert row to list, handling None values
                row_list = [str(cell) if cell is not None else "" for cell in row]
                data.append(row_list)

            df = pd.DataFrame(data)
            sheet_data[sheet_name] = df
            all_data.extend(data)

        return sheet_data, all_data

    def _prepare_context_for_gemini(self, sheet_data, max_chars=3200000):
        """
        Prepare Excel data as structured text context for Gemini.
        
        Token Limits (Gemini 2.5 Flash):
        - Input: 1,000,000 tokens (~4M characters)
        - Output: 65,536 tokens
        - Ratio: ~4 characters per token
        
        This function limits to 3.2M chars (~800K tokens = 80% of capacity)
        Uses maximum available capacity while leaving 20% safety margin.
        """
        context = "Below is data from an Excel file containing company information:\n\n"
        
        for sheet_name, df in sheet_data.items():
            sheet_context = f"=== SHEET: {sheet_name} ===\n"
            # With 3.2M char limit, we can handle very large sheets
            # No row limit - let Gemini handle as much data as possible
            sheet_str = df.to_string(index=False, max_rows=None)
            
            # Truncate if still too long
            if len(context) + len(sheet_context) + len(sheet_str) > max_chars:
                remaining_chars = max_chars - len(context) - len(sheet_context)
                if remaining_chars > 0:
                    sheet_str = sheet_str[:remaining_chars] + "\n... (truncated due to size limits)"
                else:
                    print(f"⚠️ Skipping sheet '{sheet_name}' - context limit reached")
                    break  # Skip this sheet if we're already at limit
            
            sheet_context += sheet_str
            sheet_context += "\n\n"
            context += sheet_context
        
        # Final safety check
        if len(context) > max_chars:
            context = context[:max_chars] + "\n... (truncated due to size limits)"
        
        print(f"📊 Context size: {len(context):,} chars (~{len(context)//4:,} tokens, {(len(context)//4)/10000:.1f}% of 1M limit)")
        
        return context

    def extract_with_gemini(self, model_name='gemini-2.5-flash'):
        """Extract company names from Excel using Gemini."""
        try:
            # Load workbook
            if isinstance(self.source, io.BytesIO):
                wb = openpyxl.load_workbook(self.source, data_only=True)
            else:
                file_ext = os.path.splitext(self.source)[1].lower()
                if file_ext == ".csv":
                    df = pd.read_csv(self.source)
                    buf = io.BytesIO()
                    df.to_excel(buf, index=False)
                    buf.seek(0)
                    wb = openpyxl.load_workbook(buf, data_only=True)
                else:
                    wb = openpyxl.load_workbook(self.source, data_only=True)
        except Exception as e:
            print(f"ERROR loading workbook: {e}")
            raise

        # Convert to structured text
        sheet_data, _ = self._convert_to_dataframe(wb)
        context = self._prepare_context_for_gemini(sheet_data)
        wb.close()

        # Create context-aware prompt
        if self.target_company:
            target_context = f"\nTARGET COMPANY CONTEXT: {self.target_company}\n"
            target_context += f"IMPORTANT: Use your knowledge of {self.target_company}'s industry, products, and market to filter the extracted companies.\n"
            target_context += f"Only extract companies that operate in the SAME or CLOSELY RELATED business as {self.target_company}.\n"
            target_context += f"Exclude companies from completely different industries or product categories.\n\n"
        else:
            target_context = ""

        # Create prompt for Gemini
        prompt = f"""{context}

{target_context}TASK: Extract ALL company names from the data above that are potential competitors or comparable companies.

Instructions:
- Look for columns containing company names, targets, acquirers, sellers, or similar identifiers
- Extract only actual company names (exclude headers, totals, averages, summaries)
- Ignore entries like "N/A", "TBD", "Others", "Mean", "Total", "Average", "Median"
- Include ALL companies found in the spreadsheet
- Return the results as a JSON object with the following structure:
{{
    "companies": ["Company 1", "Company 2", ...],
    "count": <number of unique companies>
}}

CRITICAL: Provide ONLY valid JSON response, no additional text, no markdown formatting, no explanations."""

        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            full_response = response.text

            # Parse JSON response
            try:
                # Remove markdown code blocks if present
                full_response = full_response.replace('```json', '').replace('```', '').strip()

                # Extract JSON from response
                json_start = full_response.find('{')
                json_end = full_response.rfind('}') + 1

                if json_start == -1 or json_end == 0:
                    return None

                json_str = full_response[json_start:json_end]
                result = json.loads(json_str)

                if "companies" not in result:
                    return None

                companies = result.get("companies", [])

                self.results = {
                    "all_companies": set(companies),
                    "total_unique": len(set(companies))
                }
                return self.results

            except json.JSONDecodeError as e:
                print(f"JSON Decode Error: {e}")
                return None

        except Exception as e:
            print(f"Exception during Gemini API call: {e}")
            return None


class CopilotResponseProcessor:
    def __init__(self, copilot_response: str, target_company: str, api_key: str):
        """
        Initialize processor with copilot response and target company name.
        """
        self.copilot_response = copilot_response
        self.target_company = target_company
        self.api_key = api_key
        self.file_paths = []
        self.relative_paths = []
        self.all_competitors = set()
        self.verified_competitors = []
        self.to_crosscheck = []
        self.file_wise_companies = {}
        
        # Configure Gemini
        genai.configure(api_key=api_key)

    def extract_file_paths(self):
        """Extract unique file paths from copilot response and cut at file extension"""
        # Pattern to match "Full Path: " followed by the path, cut at file extension
        pattern = r'Full Path:\s*(.+?\.(?:xlsx|xls|csv|pptx|pdf))'
        matches = re.findall(pattern, self.copilot_response, re.IGNORECASE)
        unique_paths = list(set(matches))
        self.file_paths = [path.strip() for path in unique_paths]

        # Extract relative paths starting from "Shared Documents/"
        self.relative_paths = []
        for path in self.file_paths:
            if "Shared Documents/" in path:
                relative_path = path.split("Shared Documents/", 1)[1]
                self.relative_paths.append(relative_path)
            else:
                self.relative_paths.append(path)

        return self.file_paths, self.relative_paths

    def classify_competitors_with_gemini(self, all_companies: list, model_name='gemini-2.5-flash'):
        """
        Use Gemini to classify extracted companies into verified competitors and to-crosscheck.
        
        Token Limits (Gemini 2.5 Flash):
        - Input: 1,000,000 tokens (~4M characters)
        - Using 80% capacity = 800K tokens available
        - This allows for ~2000 companies comfortably
        """
        companies_list = sorted(list(all_companies))
        
        # Limit to 2000 companies (80% capacity usage)
        # Each company name ~30 chars avg, 2000 companies ≈ 60K chars (~15K tokens)
        # Plus prompt (~2K tokens) = ~17K tokens total
        # Still leaves room for 783K tokens of Excel data
        if len(companies_list) > 2000:
            print(f"⚠️ Warning: {len(companies_list)} companies found. Limiting to 2000 for classification.")
            companies_list = companies_list[:2000]
        
        print(f"🔍 Classifying {len(companies_list)} companies for {self.target_company}")

        prompt = f"""You are a business analyst expert specializing in competitive analysis.

TARGET COMPANY: {self.target_company}

EXTRACTED COMPANIES CANDIDATES:
{json.dumps(companies_list, indent=2)}

TASK: Classify these candidates based on their competitive relationship with {self.target_company}.

RULES:
1. **STRICTLY** use ONLY the companies provided in the list above. DO NOT add any new companies.
2. Assign a **Confidence Score (0-100)** representing the strength of the competitive overlap.
   - 90-100: Direct competitor (same core products/services, same market).
   - 70-89: Strong competitor (significant overlap).
   - 50-69: Moderate/Indirect competitor or substitute.
   - <50: Low relevance or different industry.

CLASSIFICATION CATEGORIES:
1. **Verified Competitors**: Score >= 70. Direct/Strong competitors.
2. **To Cross-Check**: Score < 70. Indirect, potential, or unclear competitors.

RESPONSE FORMAT:
Return a JSON object with two lists. Each item must be an object containing "name" and "score".
Sort both lists by "score" in DESCENDING order.

{{
    "verified_competitors": [
        {{"name": "Company A", "score": 95, "reason": "..."}},
        {{"name": "Company B", "score": 88, "reason": "..."}}
    ],
    "to_crosscheck": [
        {{"name": "Company C", "score": 45, "reason": "..."}}
    ],
    "reasoning": "Brief analysis of the industry context."
}}"""

        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            text_response = response.text

            # Remove markdown code blocks if present
            text_response = text_response.replace('```json', '').replace('```', '').strip()

            # Extract JSON from response
            json_start = text_response.find('{')
            json_end = text_response.rfind('}') + 1
            json_str = text_response[json_start:json_end]

            result = json.loads(json_str)

            self.verified_competitors = result.get("verified_competitors", [])
            self.to_crosscheck = result.get("to_crosscheck", [])
            
            return result

        except Exception as e:
            print(f"Exception during classification: {e}")
            # Fallback: put all in to_crosscheck
            self.to_crosscheck = companies_list
            return {
                "verified_competitors": [],
                "to_crosscheck": companies_list,
                "verified_count": 0,
                "crosscheck_count": len(companies_list),
                "reasoning": "Error during classification, fallback to cross-check."
            }

def get_graph_token(tenant_id, client_id, client_secret):
    """Get Microsoft Graph API access token."""
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    try:
        app = ConfidentialClientApplication(
            client_id, authority=authority, client_credential=client_secret)
        token = app.acquire_token_for_client(
            scopes=["https://graph.microsoft.com/.default"])
        if "access_token" in token:
            return token["access_token"]
        else:
            raise Exception("No access token in response")
    except Exception as e:
        raise Exception(f"Error getting token: {e}")

def search_file_by_name(access_token, drive_id, filename):
    """Search for a file by name in SharePoint if direct path fails."""
    search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='{filename}')"
    headers = {"Authorization": f"Bearer {access_token}"}
    try:
        resp = requests.get(search_url, headers=headers)
        resp.raise_for_status()
        results = resp.json()
        if "value" in results and len(results["value"]) > 0:
            return results["value"][0]
        return None
    except Exception:
        return None

def download_file_from_sharepoint(access_token, drive_id, relative_path):
    """Download file from SharePoint using Microsoft Graph API with fallback search."""
    # Try direct path first
    encoded_path = quote(relative_path, safe='/')
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded_path}:/content"
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        resp = requests.get(url, headers=headers)

        if resp.status_code == 404:
            filename = os.path.basename(relative_path)
            file_item = search_file_by_name(access_token, drive_id, filename)

            if file_item:
                download_url = file_item.get('@microsoft.graph.downloadUrl')
                if download_url:
                    resp = requests.get(download_url)
                    resp.raise_for_status()
                    return io.BytesIO(resp.content)
                else:
                    raise Exception(f"No download URL for file: {filename}")
            else:
                raise Exception(f"File not found: {filename}")

        resp.raise_for_status()
        return io.BytesIO(resp.content)

    except Exception as e:
        raise Exception(f"Download failed: {e}")
